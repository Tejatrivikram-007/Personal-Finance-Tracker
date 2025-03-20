from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm, PasswordResetForm
from django.contrib import messages
from django.contrib.auth import login, logout, password_validation
from django.contrib.auth import update_session_auth_hash
from .middlewares import auth, guest
from .models import Income, Expense,Savings
from .forms import IncomeForm, ExpenseForm,SavingsForm
import csv
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side


#  -------------------------<< Register >>----------------------
@guest
def register_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        initial_data = {'username':'', 'password1':'','password2':""}
        form = UserCreationForm(initial=initial_data)
    return render(request, 'auth/register.html',{'form':form})

#  -------------------------<< Login >>----------------------
@guest
def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request,user)
            return redirect('dashboard')
    else:
        initial_data = {'username':'', 'password':''}
        form = AuthenticationForm(initial=initial_data)
    return render(request, 'auth/login.html',{'form':form}) 

#  -------------------------<< Reset Password----------------

@guest
def reset_password_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        
        if password1 != password2:
            messages.error(request, "Passwords do not match.")
            return redirect('reset_password')

        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            messages.error(request, "Username does not exist.")
            print("Username does not exist.")
            return redirect('reset_password')

        try:
            password_validation.validate_password(password1, user)  # Validate the password
        except Exception as e:
            messages.error(request, f"Password error: {e}")
            print(f"{e}")
            return redirect('reset_password')
                
        user.set_password(password1)   # Set the new password
        print(user.set_password(password1))
        user.save()

        update_session_auth_hash(request, user) # Update the session so the user doesn't get logged out

        messages.success(request, "Your password has been successfully updated.")

    # GET request: Render the reset password form
    return render(request, 'auth/reset_password.html')


#  -------------------------<< Logout >>----------------------
def logout_view(request):
    logout(request)
    return redirect('login')


#  -------------------------<< MAIN >>----------------------

@auth
def dashboard(request):
    incomes = Income.objects.filter(user=request.user)
    expenses = Expense.objects.filter(user=request.user)
    savings = Savings.objects.filter(user=request.user)
    total_income = sum(income.amount for income in incomes)
    total_expense = sum(expense.amount for expense in expenses)
    total_savings=sum(saving.amount for saving in savings)
    balance = total_income - (total_expense + total_savings)
    
    # Debugging statements
    print(f"Total Income: {total_income}, Total Expense: {total_expense}, Balance: {balance}")

    return render(request, 'pages/dashboard.html', {
        'incomes': incomes,
        'expenses': expenses,
        'savings': savings,
        'total_income': total_income,
        'total_expense': total_expense,
        'total_savings': total_savings,
        'balance': balance,
    })

@auth
def finance(request):
    incomes = Income.objects.filter(user=request.user).order_by('date')
    expenses = Expense.objects.filter(user=request.user).order_by('date')
    savings= Savings.objects.filter(user=request.user).order_by('date')
    return render(request, 'pages/finance.html', {
        'incomes': incomes,
        'expenses': expenses,
        'savings':savings
        })


def add_income(request):
    if request.method == 'POST':
        form = IncomeForm(request.POST)
        if form.is_valid():
            income = form.save(commit=False)
            income.user = request.user
            income.save()
            return redirect('dashboard')
    else:
        form = IncomeForm()
    return render(request, 'pages/add_income.html', {'form': form})


def add_expense(request):
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user
            expense.save()
            return redirect('dashboard')
    else:
        form = ExpenseForm()
    return render(request, 'pages/add_expense.html', {'form': form})

def add_savings(request):
    if request.method == 'POST':
        form = SavingsForm(request.POST)
        if form.is_valid():
            savings = form.save(commit=False)
            savings.user = request.user
            savings.save()
            return redirect('dashboard')
    else:
        form = SavingsForm()
    return render(request, 'pages/add_savings.html', {'form': form})


def update_expense(request, id):
    expense = Expense.objects.get(id=id, user=request.user)
    expenses = get_object_or_404(Expense, id=id, user=request.user)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense) 

        if form.is_valid():  
            form.save()  
            return redirect('finance')  
    else:
        form = ExpenseForm(instance=expense)  

    return render(request, 'pages/update_expense.html', {'form': form, 'expense': expenses})

def delete_expense(request,pk):
    try:
        expense=Expense.objects.get(id=pk,user=request.user)
    except:
        return HttpResponse(f'The expense with id= {pk} does not exist')
    expense.delete()
    return redirect('finance')


def update_income(request, id):
    income = Income.objects.get(id=id, user=request.user)
    incomes = get_object_or_404(Income, id=id, user=request.user)
    if request.method == 'POST':
        form = IncomeForm(request.POST, instance=income) 

        if form.is_valid():  
            form.save()  
            return redirect('finance')  
    else:
        form = IncomeForm(instance=income)  

    return render(request, 'pages/update_income.html', {'form': form, 'expense': income})

def delete_income(request,pk):
    try:
        income=Income.objects.get(id=pk,user=request.user)
    except:
        return HttpResponse(f'The expense with id= {pk} does not exist')
    income.delete()
    return redirect('finance')

def update_savings(request, id):
    savings = Savings.objects.get(id=id, user=request.user)
    savingses = get_object_or_404(Savings, id=id, user=request.user)
    if request.method == 'POST':
        form = SavingsForm(request.POST, instance=savings)
        if form.is_valid():
            form.save()
            return redirect('finance')
    else:
        form = SavingsForm(instance=savings)

    return render(request, 'pages/update_savings.html', {'form': form, 'savings':savings})

def delete_savings(request,pk):
    try:
        savings=Savings.objects.get(id=pk,user=request.user)
    except:
        return HttpResponse(f'The savings with id= {pk} does not exist')
    savings.delete()
    return redirect('finance')


@auth
def statements(request):
   
    year_range = range(2024, 2036)
    context = {
        'year_range': year_range,
    }
    return render(request, 'pages/statements_form.html', context)



@auth
def generate_statement(request, month, year):
    # Retrieve the income and expense records for the selected month and year, ordered by date
    income_records = Income.objects.filter(user=request.user, date__month=month, date__year=year).order_by('date')
    expense_records = Expense.objects.filter(user=request.user, date__month=month, date__year=year).order_by('date')
    savings_records = Savings.objects.filter(user=request.user, date__month=month, date__year= year).order_by('date')
    total_income = sum(income.amount for income in income_records)
    total_expense = sum(expense.amount for expense in expense_records)
    total_savings = sum(savings.amount for savings in savings_records)
    balance = total_income - total_expense

    download_option = request.GET.get('download')

    if download_option == 'csv':
        return download_statement_csv(income_records, expense_records,savings_records, month, year, total_income, total_expense,total_savings, balance)
    elif download_option == 'pdf':
        return download_statement_pdf(income_records, expense_records,savings_records, month, year, total_income, total_expense,total_savings, balance)
    elif download_option == 'excel':
        return download_statement_excel(income_records, expense_records, savings_records, month, year, total_income, total_expense, total_savings, balance)

    context = {
        'month': month,
        'year': year,
        'income_records': income_records,
        'expense_records': expense_records,
        'savings_records': savings_records,
        'total_income': total_income,
        'total_expense': total_expense,
        'total_savings': total_savings,
        'balance': balance,
    }

    return render(request, 'pages/statement.html', context)


def download_statement_csv(income_records, expense_records, savings_records, month, year, total_income, total_expense, total_savings, balance):
   
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="statement_{year}_{month}.csv"'  # download filename

    # Create a CSV writer
    writer = csv.writer(response)

    # Title of the CSV
    writer.writerow([f'Statement for {month}/{year}'])  # Dynamic Title for the month and year
    writer.writerow([])  # Add an empty row after the title for better spacing

    # Income Records Section
    writer.writerow(['Income Records'])  # Title for Income section
    writer.writerow(['Date', 'Source', 'Description', 'Amount'])  # Header row for Income records

    # Write income records to the CSV, ordered by date
    for income in income_records:
        writer.writerow([income.date, income.source, income.description, income.amount])

    writer.writerow([])  # Add an empty row between sections

    # Expense Records Section
    writer.writerow(['Expense Records'])  # Title for Expense section
    writer.writerow(['Date', 'Category', 'Description', 'Amount'])  # Header row for Expense records

    # Write expense records to the CSV, ordered by date
    for expense in expense_records:
        writer.writerow([expense.date, expense.category, expense.description, expense.amount])

    writer.writerow([])  # Add an empty row between sections

    # Savings Records Section
    writer.writerow(['Savings Records'])  # Title for Savings section
    writer.writerow(['Date', 'Category', 'Description', 'Amount'])  # Header row for Savings records

    # Write savings records to the CSV, ordered by date
    for saving in savings_records:
        writer.writerow([saving.date, saving.category, saving.description, saving.amount])

    writer.writerow([])  # Add an empty row between sections

    # Add the totals and balance at the bottom
    writer.writerow(['Total Income', total_income])
    writer.writerow(['Total Expense', total_expense])
    writer.writerow(['Total Savings', total_savings])
    writer.writerow(['Balance', balance])

    return response


def download_statement_pdf(income_records, expense_records, savings_records, month, year, total_income, total_expense, total_savings, balance):
    # Create a PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="statement_{year}_{month}.pdf"'

    # Create a PDF document using ReportLab
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Title
    p.setFont('Helvetica-Bold', 16)
    p.drawString(200, height - 40, f"Statement for {month}/{year}")

    # Income Records Header
    p.setFont('Helvetica-Bold', 12)
    p.drawString(50, height - 80, 'Income Records:')
    p.setFont('Helvetica', 10)
    p.drawString(50, height - 100, 'Date')
    p.drawString(150, height - 100, 'Source')
    p.drawString(250, height - 100, 'Description')
    p.drawString(350, height - 100, 'Amount')

    # Income Records Data (ordered by date)
    y_position = height - 120
    for income in income_records:
        p.drawString(50, y_position, str(income.date))
        p.drawString(150, y_position, income.source)
        p.drawString(250, y_position, income.description)
        p.drawString(350, y_position, str(income.amount))
        y_position -= 20

    # Add space between sections
    y_position -= 20

    # Expense Records Header
    p.setFont('Helvetica-Bold', 12)
    p.drawString(50, y_position, 'Expense Records:')
    p.setFont('Helvetica', 10)
    p.drawString(50, y_position - 20, 'Date')
    p.drawString(150, y_position - 20, 'Category')
    p.drawString(250, y_position - 20, 'Description')
    p.drawString(350, y_position - 20, 'Amount')

    # Expense Records Data (ordered by date)
    y_position -= 40
    for expense in expense_records:
        p.drawString(50, y_position, str(expense.date))
        p.drawString(150, y_position, expense.category)
        p.drawString(250, y_position, expense.description)
        p.drawString(350, y_position, str(expense.amount))
        y_position -= 20

    # Add space between sections
    y_position -= 20

    # Savings Records Header
    p.setFont('Helvetica-Bold', 12)
    p.drawString(50, y_position, 'Savings Records:')
    p.setFont('Helvetica', 10)
    p.drawString(50, y_position - 20, 'Date')
    p.drawString(150, y_position - 20, 'Source')
    p.drawString(250, y_position - 20, 'Description')
    p.drawString(350, y_position - 20, 'Amount')

    # Savings Records Data (ordered by date)
    y_position -= 40
    for savings in savings_records:
        p.drawString(50, y_position, str(savings.date))
        p.drawString(150, y_position, savings.category)
        p.drawString(250, y_position, savings.description)
        p.drawString(350, y_position, str(savings.amount))
        y_position -= 20

    # Add totals and balance at the bottom
    y_position -= 20
    p.setFont('Helvetica-Bold', 12)
    p.drawString(50, y_position, f"Total Income: {total_income}")
    p.drawString(50, y_position - 20, f"Total Expense: {total_expense}")
    p.drawString(50, y_position - 40, f"Total Savings: {total_savings}")
    p.drawString(50, y_position - 60, f"Balance: {balance}")

    # Finalize PDF
    p.showPage()
    p.save()

    return response


def download_statement_excel(income_records, expense_records, savings_records, month, year, total_income, total_expense, total_savings, balance):
    # Create an Excel workbook and a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Statement {month}-{year}"

    # Set the font and alignment for the title
    title = f'Statement for {month}/{year}'
    ws.merge_cells('A1:F1')  # Merge the title across columns
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set the font for subheadings and headers
    header_font = Font(size=12, bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Add Income Records
    row = 3
    ws.cell(row=row, column=1).value = 'Income Records'
    ws.cell(row=row, column=1).font = Font(size=14, bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    row += 1

    # Table headers for Income
    headers = ['Date', 'Source', 'Description', 'Amount']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(bottom=Side(style='thin'))

    row += 1

    # Add income records data
    for income in income_records:
        ws.cell(row=row, column=1).value = income.date
        ws.cell(row=row, column=2).value = income.source
        ws.cell(row=row, column=3).value = income.description
        ws.cell(row=row, column=4).value = income.amount
        row += 1

    # Add an empty row between sections
    row += 1

    # Add Expense Records
    ws.cell(row=row, column=1).value = 'Expense Records'
    ws.cell(row=row, column=1).font = Font(size=14, bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    row += 1

    # Table headers for Expense
    headers = ['Date', 'Category', 'Description', 'Amount']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(bottom=Side(style='thin'))

    row += 1

    # Add expense records data
    for expense in expense_records:
        ws.cell(row=row, column=1).value = expense.date
        ws.cell(row=row, column=2).value = expense.category
        ws.cell(row=row, column=3).value = expense.description
        ws.cell(row=row, column=4).value = expense.amount
        row += 1

    # Add an empty row between sections
    row += 1

    # Add Savings Records
    ws.cell(row=row, column=1).value = 'Savings Records'
    ws.cell(row=row, column=1).font = Font(size=14, bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    row += 1

    # Table headers for Savings
    headers = ['Date', 'Category', 'Description', 'Amount']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(bottom=Side(style='thin'))

    row += 1

    # Add savings records data
    for saving in savings_records:
        ws.cell(row=row, column=1).value = saving.date
        ws.cell(row=row, column=2).value = saving.category
        ws.cell(row=row, column=3).value = saving.description
        ws.cell(row=row, column=4).value = saving.amount
        row += 1

    # Add an empty row between sections
    row += 1

    # Add the totals and balance
    totals = [
        ('Total Income', total_income),
        ('Total Expense', total_expense),
        ('Total Savings', total_savings),
        ('Balance', balance)
    ]

    for total_name, total_value in totals:
        ws.cell(row=row, column=1).value = total_name
        ws.cell(row=row, column=2).value = total_value
        row += 1

    # Create the HTTP response to serve the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="statement_{year}_{month}.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response
